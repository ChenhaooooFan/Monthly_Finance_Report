"""
ColorFour LLC — 财务流水智能分类系统
单文件版 | 支持 BofA / Chase PDF + Payoneer CSV
"""
import re
import io
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════
#  RULES
# ══════════════════════════════════════════════════════════════
PAYONEER_RULES = {
    "5133":"员工工资（派安盈）","9382":"员工工资（派安盈）","1510":"员工工资（派安盈）",
    "0076":"员工工资（派安盈）","2000":"员工工资（派安盈）","0472":"员工工资（派安盈）",
    "1591":"员工工资（派安盈）","7537":"员工工资（派安盈）","4745":"员工工资（派安盈）",
    "0347":"员工工资（派安盈）","6721":"穿戴甲进货成本","4475":"穿戴甲进货成本",
    "6667":"穿戴甲进货成本","5675":"穿戴甲进货成本","0528":"穿戴甲进货成本",
    "9491":"穿戴甲进货成本","5318":"穿戴甲进货成本","9276":"穿戴甲进货成本",
    "3284":"穿戴甲进货成本","4783":"穿戴甲进货成本","9231":"穿戴甲进货成本",
    "8436":"穿戴甲进货成本","0227":"穿戴甲进货成本","3121":"穿戴甲进货成本",
    "5755":"穿戴甲进货成本","9460":"穿戴甲进货成本","4785":"物流成本",
    "0664":"物流成本","2292":"物流成本","2465":"配套耗材进货成本",
    "8011":"配套耗材进货成本","9005":"配套耗材进货成本","8456":"配套耗材进货成本",
    "6006":"拍摄费用","2689":"拍摄费用","9215":"拍摄费用","6872":"拍摄费用",
    "1530":"办公软件会费","1001":"办公软件会费",
}

BANK_RULES = [
    (r"Zelle.*JINGYI HUANG",                    "员工工资（现金）"),
    (r"Zelle.*ZHENLI WANG",                     "国内转账"),
    (r"Zelle.*(InternalTransfer|COLORFOUR LLC)", "Payoneer内部转账"),
    (r"Zelle.*Lunch",                            "午餐补助费"),
    (r"Zelle.*Parking",                          "停车补助费"),
    (r"Zelle.*TINGTING LIN",                     "午餐补助费"),
    (r"Zelle.*payroll",                          "主播工资"),
    (r"Zelle payment to [A-Z][a-z]",             "主播工资"),
    (r"Zelle Payment To|Zelle Payment From",     "朋友间转账"),
    (r"PURCHASE REFUND.*TIKTOK",                 "销售退款冲销"),
    (r"TIKTOK ADS|TIKTOK\.COM CA",               "广告费 (TK Ads)"),
    (r"TikTok Inc.*PAYMENT|TikTok Shop.*DES:",   "TikTok净销售额"),
    (r"SHOPIFY\*",                               "Shopify扣款"),
    (r"ADP WAGE PAY",                            "员工工资"),
    (r"ADP Tax",                                 "APD TAX"),
    (r"ADP PAY-BY-PAY",                          "办公软件会费"),
    (r"ADP PAYROLL FEES",                        "行政费用"),
    (r"Regus Management",                        "房租成本"),
    (r"H\.K ALADDIN",                            "穿戴甲进货成本"),
    (r"Shantou Panjia",                          "配套耗材进货成本"),
    (r"Mars Shipping",                           "物流成本"),
    (r"HANKIN PATENT LAW",                       "律师费用"),
    (r"OPENAI|KLAVIYO|GOOGLE.*Worksp|INTUIT.*QBooks|TIGER-ROAR|imyfone", "办公软件会费"),
    (r"TRACK1099",                               "行政费用"),
    (r"AMAZON MKTPL|AMAZON RETA",                "办公用品采购"),
    (r"FANTUAN|HUNGRYPANDA|UBER.*EATS|SQ.*DOWNTOWN LA PHO|RUSSELL.*CONVENIENCE|SLICES.*LOS ANGELES", "午餐补助费"),
    (r"Tst\* Salata|Sq \*Hinodeya|Uep\*Lao|Tst\* Togo|King Charcoal|Hanshin|LA Tofu|Lao MA Tou|Samyi Cake|Angel.*Tacos|Song Yu|Xiaoqing Huang|M&D Snacks|Full House|Tianlu Investment|Groupon|Classpass|GOLDEN DRAGON", "午餐补助费"),
    (r"99 Ranch|H Mart|Gelson|Ralphs|168 Market|Super King|Home Depot|Daiso", "办公用品采购"),
    (r"TRANSFER.*ColorFour LLC Confirmation|TRANSFER COLORFOUR LLC:ColorFour", "Payoneer内部转账"),
    (r"External transfer fee|Monthly Service Fee", "银行手续费"),
    (r"Check #1[0-9]+",                          "停车补助费"),
    (r"PAYPAL.*INST XFER",                       "其他支出"),
]

ALL_CATS = [
    "广告费 (TK Ads)","穿戴甲进货成本","配套耗材进货成本","物流成本",
    "员工工资","员工工资（派安盈）","员工工资（现金）","主播工资",
    "拍摄费用","房租成本","办公软件会费","办公用品采购","行政费用",
    "银行手续费","APD TAX","Shopify扣款","停车补助费","午餐补助费",
    "国内转账","销售税预提(CA)","律师费用","TikTok净销售额","Shopify净销售额",
    "销售退款冲销","Payoneer内部转账","朋友间转账","其他支出","待确认",
]

EXCLUDE_PL = {"Payoneer内部转账", "销售退款冲销", "朋友间转账"}

EXPENSE_ORDER = [
    ("广告费（TK Ads）",      ["广告费 (TK Ads)"]),
    ("物流成本",              ["物流成本"]),
    ("穿戴甲进货成本",         ["穿戴甲进货成本"]),
    ("配套耗材进货成本",       ["配套耗材进货成本"]),
    ("停车补助费",            ["停车补助费"]),
    ("午餐补助费",            ["午餐补助费"]),
    ("房租成本",              ["房租成本"]),
    ("行政费用",              ["行政费用"]),
    ("办公用品采购",           ["办公用品采购"]),
    ("办公软件会员费用",       ["办公软件会费"]),
    ("银行手续费用",           ["银行手续费"]),
    ("员工工资（Zelle/现金）", ["员工工资","员工工资（现金）"]),
    ("拍摄费用",              ["拍摄费用"]),
    ("主播工资",              ["主播工资"]),
    ("国内转账",              ["国内转账"]),
    ("Shopify 扣款",          ["Shopify扣款"]),
    ("员工工资（派安盈）",     ["员工工资（派安盈）"]),
    ("APD TAX",               ["APD TAX"]),
    ("律师费用",              ["律师费用"]),
    ("销售税预提（CA）",       ["销售税预提(CA)"]),
    ("其他支出",              ["其他支出"]),
]


def cls_bank(desc):
    for pat, cat in BANK_RULES:
        if re.search(pat, desc, re.IGNORECASE):
            return cat
    return "待确认"


def cls_pay(desc):
    if re.search(r"Payment from COLORFOUR LLC", desc, re.IGNORECASE):
        return "Payoneer内部转账"
    if re.search(r"1688\.com", desc, re.IGNORECASE):
        return "办公用品采购"
    m = re.search(r"\((\d{4})\)", desc)
    if m and m.group(1) in PAYONEER_RULES:
        return PAYONEER_RULES[m.group(1)]
    return "待确认"


# ══════════════════════════════════════════════════════════════
#  PARSERS
# ══════════════════════════════════════════════════════════════
def parse_amount(s):
    try:
        return float(str(s).replace(",", "").replace("$", "").strip())
    except Exception:
        return 0.0


def detect_bank(text):
    if "Bank of America" in text or "bankofamerica" in text.lower():
        return "BofA"
    if "Chase" in text or "JPMorgan" in text:
        return "Chase"
    return "BofA"


def parse_bofa(data: bytes) -> pd.DataFrame:
    try:
        import pdfplumber
    except ImportError:
        st.error("缺少依赖: pip install pdfplumber")
        return pd.DataFrame()

    with pdfplumber.open(io.BytesIO(data)) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    SKIP_PREFIXES = (
        "Date Description", "continued on the next page", "BUSINESS ADVANTAGE",
        "Your checking account", "Page ", "COLORFOUR LLC !", "See the big picture",
        "To learn more", "When you use the QRC", "Mobile Banking requires",
        "Message and data rates", "Help prevent", "Consider writing",
        "You can also set up", "Scan the code", "Moving from checks",
        "Please see", "SSM", "PULL:", "NEW:", "Find more", "Explore your",
        "Bank of America", "Equal Housing", "bofa.com", "bankofamerica.com",
    )

    DATE_LINE  = re.compile(r'^(\d{2}/\d{2})/\d{2}\s+(.+?)\s+(-?[\d,]+\.\d{2})\s*$')
    CHECK_LINE = re.compile(r'^(\d{2}/\d{2})/\d{2}\s+(\d+)\s+-?([\d,]+\.\d{2})\s*$')

    def skip(line):
        if not line: return True
        for p in SKIP_PREFIXES:
            if line.startswith(p): return True
        return False

    lines = text.split("\n")
    rows = []
    section = None
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        if line in ("Deposits and other credits", "Deposits and other credits - continued"):
            section = "dep"; i += 1; continue
        if line in ("Withdrawals and other debits", "Withdrawals and other debits - continued"):
            section = "wdl"; i += 1; continue
        if line.startswith("Checks") and "Total" not in line and "image" not in line:
            section = "chk"; i += 1; continue
        if "Service fees" in line and "Total" not in line and len(line) < 20:
            section = "fee"; i += 1; continue
        if any(x in line for x in ["Daily ledger", "Check images",
               "Total deposits", "Total withdrawals", "Total checks",
               "Total service", "Total # of", "Note your Ending"]):
            i += 1; continue

        if section in ("dep", "wdl", "chk", "fee") and re.match(r'^\d{2}/\d{2}/\d{2}', line):
            j = i + 1
            m_first = DATE_LINE.match(line) if section != "chk" else None
            if m_first:
                full = line
            else:
                parts = [line]
                while j < len(lines):
                    nxt = lines[j].strip()
                    if not nxt or re.match(r'^\d{2}/\d{2}/\d{2}', nxt) or skip(nxt): break
                    if nxt.startswith("Total"): break    
                    if nxt in ("Deposits and other credits",
                               "Deposits and other credits - continued",
                               "Withdrawals and other debits",
                               "Withdrawals and other debits - continued",
                               "Checks", "Service fees"): break
                    parts.append(nxt)
                    j += 1
                full = " ".join(parts)

            if section == "chk":
                m = CHECK_LINE.match(full)
                if m:
                    rows.append({"date": m.group(1), "desc": f"Check #{m.group(2)}",
                                 "amount": -parse_amount(m.group(3)), "src": "BofA"})
                    i = j; continue
            else:
                m = DATE_LINE.match(full)
                if m:
                    amt = parse_amount(m.group(3))
                    if section in ("wdl", "fee") and amt > 0:
                        amt = -amt
                    rows.append({"date": m.group(1), "desc": m.group(2).strip(),
                                 "amount": amt, "src": "BofA"})
                    i = j; continue
        i += 1

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["cat"]    = df["desc"].apply(cls_bank)
    df["status"] = df["cat"].apply(lambda c: "auto" if c != "待确认" else "pending")
    return df


def parse_chase(data: bytes) -> pd.DataFrame:
    try:
        import pdfplumber
    except ImportError:
        st.error("缺少依赖: pip install pdfplumber")
        return pd.DataFrame()

    rows = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        text = "\n".join(p.extract_text() or "" for p in pdf.pages)

    section = None
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        lu = line.upper()
        if "DEPOSITS AND ADDITIONS" in lu:
            section = "dep"
            continue
        if "ATM & DEBIT CARD WITHDRAWALS" in lu and "SUMMARY" not in lu:
            section = "wdl"
            continue
        if "ELECTRONIC WITHDRAWALS" in lu:
            section = "ewdl"
            continue
        if lu.startswith("FEES"):
            section = "fee"
            continue
        if "DAILY ENDING BALANCE" in lu or "ATM & DEBIT CARD SUMMARY" in lu:
            section = None
            continue

        if section in ("dep", "wdl", "ewdl", "fee"):
            m = re.match(r"^(\d{2}/\d{2})\s+(.+?)\s+\$?([\d,]+\.\d{2})$", line)
            if not m:
                m = re.match(r"^(\d{2}/\d{2})\s+(.+?)\s+([\d,]+\.\d{2})$", line)
            if m:
                desc = m.group(2).strip()
                if any(x in desc for x in ["Total", "Beginning Balance", "Ending Balance"]):
                    continue
                amt = parse_amount(m.group(3))
                amount = amt if section == "dep" else -amt
                rows.append({"date": m.group(1), "desc": desc, "amount": amount, "src": "Chase"})

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df["cat"]    = df["desc"].apply(cls_bank)
    df["status"] = df["cat"].apply(lambda c: "auto" if c != "待确认" else "pending")
    return df


def parse_payoneer(data: bytes) -> pd.DataFrame:
    import datetime
    try:
        df_raw = pd.read_csv(io.BytesIO(data), encoding="utf-8-sig")
    except Exception:
        df_raw = pd.read_csv(io.BytesIO(data), encoding="gbk")
    df_raw.columns = [c.strip() for c in df_raw.columns]
    date_col = next((c for c in df_raw.columns if "date" in c.lower()), None)
    desc_col = next((c for c in df_raw.columns if "desc" in c.lower()), None)
    amt_col  = next((c for c in df_raw.columns if "amount" in c.lower()), None)
    if not all([date_col, desc_col, amt_col]):
        return pd.DataFrame()
    rows = []
    for _, row in df_raw.iterrows():
        try:
            amount = float(str(row[amt_col]).replace(",","").replace("$","").strip())
        except Exception:
            continue
        date_s = str(row[date_col]).strip()
        for fmt in ["%d %b, %Y","%Y-%m-%d","%m/%d/%Y","%d/%m/%Y"]:
            try:
                date_s = datetime.datetime.strptime(date_s, fmt).strftime("%m/%d")
                break
            except Exception:
                pass
        desc = str(row[desc_col]).strip()
        cat  = cls_pay(desc)
        rows.append({"date": date_s, "desc": desc, "amount": amount,
                     "cat": cat, "src": "Payoneer",
                     "status": "auto" if cat != "待确认" else "pending"})
    return pd.DataFrame(rows)


def extract_bofa_summary(data: bytes) -> dict:
    """Extract cover page totals from BofA PDF for reconciliation."""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            text = pdf.pages[0].extract_text() or ""
    except Exception:
        return {}
    result = {}
    patterns = {
        "deposits":    r"Deposits and other credits\s+([\d,]+\.\d{2})",
        "withdrawals": r"Withdrawals and other debits\s+-?([\d,]+\.\d{2})",
        "checks":      r"Checks\s+-?([\d,]+\.\d{2})",
        "fees":        r"Service fees\s+-?([\d,]+\.\d{2})",
        "ending":      r"Ending balance[^\$]*\$([\d,]+\.\d{2})",
    }
    for key, pat in patterns.items():
        m = re.search(pat, text)
        if m:
            result[key] = float(m.group(1).replace(",",""))
    return result


def extract_chase_summary(data: bytes) -> dict:
    """Extract cover page totals from Chase PDF for reconciliation."""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            text = pdf.pages[0].extract_text() or ""
    except Exception:
        return {}
    result = {}
    patterns = {
        "deposits":    r"Total Deposits and Additions\s+\$?([\d,]+\.\d{2})",
        "withdrawals": r"(?:ATM & Debit Card Withdrawals|Total ATM)[^\d]*([\d,]+\.\d{2})",
        "ending":      r"Ending Balance\s+\d+\s+\$?([\d,]+\.\d{2})",
    }
    for key, pat in patterns.items():
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            result[key] = float(m.group(1).replace(",",""))
    return result


def parse_bank_pdf(data: bytes):
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            first = pdf.pages[0].extract_text() or ""
    except Exception:
        first = ""
    bank = detect_bank(first)
    if bank == "Chase":
        df = parse_chase(data)
        summary = extract_chase_summary(data)
        return df, "Chase", summary
    df = parse_bofa(data)
    summary = extract_bofa_summary(data)
    return df, "BofA", summary


# ══════════════════════════════════════════════════════════════
#  SUMMARY & EXPORT
# ══════════════════════════════════════════════════════════════
def build_summary(df, shopify=0.0):
    totals = {}
    for _, r in df.iterrows():
        totals[r["cat"]] = totals.get(r["cat"], 0.0) + r["amount"]
    tk  = totals.get("TikTok净销售额", 0.0)
    inc = tk + shopify
    exp = sum(v for k, v in totals.items() if v < 0 and k not in EXCLUDE_PL)
    net = inc + exp
    return {"totals": totals, "tk": tk, "shopify": shopify,
            "income": inc, "expense": exp, "net": net,
            "margin": round(net/inc*100,1) if inc else 0}


def export_csv(df, month, shopify=0.0):
    s = build_summary(df, shopify)
    t = s["totals"]
    lines = ["\ufeff", f"{month} 财务报表 — 费用支出明细\n",
             "序号,支出项目名称,金额,备注\n"]
    total_exp = 0.0
    for i,(label,keys) in enumerate(EXPENSE_ORDER,1):
        amt = sum(abs(t.get(k,0)) for k in keys)
        total_exp += amt
        lines.append(f'{i},"{label}","${amt:,.2f}",\n')
    lines.append(f',"支出合计","${total_exp:,.2f}",\n\n')
    lines += [f"{month} 财务报表 — 销售收入明细\n",
              "序号,收入项目名称,金额,备注\n",
              f'1,"TikTok净销售额","${s["tk"]:,.2f}","退款冲销已扣除"\n',
              f'2,"Shopify净销售额","${s["shopify"]:,.2f}",{"请手动填入" if shopify==0 else ""}\n',
              f'3,"亚马逊退货净额","$0.00",\n',
              f'4,"USPS退款净额","$0.00",\n',
              f',"收入合计","${s["income"]:,.2f}",\n\n',
              f'"净利润","${s["net"]:,.2f}"\n',
              f'"净利润率","{s["margin"]}%"\n\n',
              "完整流水分类明细\n",
              "日期,来源,描述,金额,分类,状态\n"]
    for _,r in df.iterrows():
        desc = str(r["desc"]).replace('"','""')
        lines.append(f'"{r["date"]}","{r["src"]}","{desc}","{r["amount"]:,.2f}","{r["cat"]}","{r["status"]}"\n')
    return "".join(lines).encode("utf-8")


def export_excel(df, month, shopify=0.0):
    s  = build_summary(df, shopify)
    t  = s["totals"]
    MID,DARK,LITE,WHITE = "2E75B6","1F4E79","D6E4F0","FFFFFF"
    def sd(c="BDD7EE"):
        x=Side(style="thin",color=c); return Border(left=x,right=x,top=x,bottom=x)
    def hdr(c,v,bg=MID):
        c.value=v; c.font=Font(name="Arial",bold=True,color=WHITE,size=10)
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=sd()
    def dat(c,v,bg=WHITE,bold=False,align="left",color="000000",fmt=None):
        c.value=v; c.font=Font(name="Arial",bold=bold,size=10,color=color)
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal=align,vertical="center")
        c.border=sd()
        if fmt: c.number_format=fmt

    wb = Workbook()
    ws=wb.active; ws.title="支出明细"
    for col,w in zip("ABCD",[6,24,14,16]): ws.column_dimensions[col].width=w
    ws.merge_cells("A1:D1"); ws.row_dimensions[1].height=34
    c=ws["A1"]; c.value=f"{month} 费用支出明细表"
    c.font=Font(name="Arial",bold=True,size=14,color=WHITE)
    c.fill=PatternFill("solid",fgColor=MID)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=sd()
    ws.row_dimensions[2].height=22
    for col,txt in enumerate(["序号","支出项目名称","金额","备注"],1): hdr(ws.cell(2,col),txt)
    total_exp=0.0
    for i,(label,keys) in enumerate(EXPENSE_ORDER):
        r=i+3; ws.row_dimensions[r].height=20
        bg=LITE if i%2==0 else WHITE
        amt=sum(abs(t.get(k,0)) for k in keys); total_exp+=amt
        dat(ws.cell(r,1),i+1,bg=bg,align="center")
        dat(ws.cell(r,2),label,bg=bg)
        c=ws.cell(r,3); c.value=amt; c.font=Font(name="Arial",size=10)
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.number_format='$#,##0.00'; c.border=sd()
        dat(ws.cell(r,4),"",bg=bg)
    rT=len(EXPENSE_ORDER)+3; ws.merge_cells(f"A{rT}:B{rT}"); ws.row_dimensions[rT].height=24
    hdr(ws.cell(rT,1),"支出合计",bg=DARK)
    ws.cell(rT,2).fill=PatternFill("solid",fgColor=DARK); ws.cell(rT,2).border=sd()
    c=ws.cell(rT,3); c.value=total_exp
    c.font=Font(name="Arial",bold=True,size=11,color=WHITE)
    c.fill=PatternFill("solid",fgColor=DARK)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.number_format='$#,##0.00'; c.border=sd()
    hdr(ws.cell(rT,4),"",bg=DARK)

    ws2=wb.create_sheet("收入明细")
    for col,w in zip("ABCD",[6,24,14,20]): ws2.column_dimensions[col].width=w
    ws2.merge_cells("A1:D1"); ws2.row_dimensions[1].height=34
    c2=ws2["A1"]; c2.value=f"{month} 销售收入明细表"
    c2.font=Font(name="Arial",bold=True,size=14,color=WHITE)
    c2.fill=PatternFill("solid",fgColor=MID)
    c2.alignment=Alignment(horizontal="center",vertical="center"); c2.border=sd()
    ws2.row_dimensions[2].height=22
    for col,txt in enumerate(["序号","收入项目名称","金额","备注"],1): hdr(ws2.cell(2,col),txt)
    income_rows=[("TikTok净销售额",s["tk"],"退款冲销已扣除"),
                 ("Shopify净销售额",s["shopify"],"需手动填入" if shopify==0 else ""),
                 ("亚马逊退货净额",0.0,""),("USPS退款净额",0.0,""),("房租退款净额",0.0,"")]
    for i,(label,amt,note) in enumerate(income_rows):
        r=i+3; ws2.row_dimensions[r].height=20
        bg=LITE if i%2==0 else WHITE
        dat(ws2.cell(r,1),i+1,bg=bg,align="center")
        dat(ws2.cell(r,2),label,bg=bg)
        c=ws2.cell(r,3); c.value=amt if amt else None
        c.font=Font(name="Arial",size=10)
        c.fill=PatternFill("solid",fgColor="FFF2CC" if "手动" in note else bg)
        c.alignment=Alignment(horizontal="center",vertical="center")
        c.number_format='$#,##0.00'; c.border=sd()
        dat(ws2.cell(r,4),note,bg=bg,color="856404" if "手动" in note else "000000")
    rI=len(income_rows)+3; ws2.merge_cells(f"A{rI}:B{rI}"); ws2.row_dimensions[rI].height=24
    hdr(ws2.cell(rI,1),"收入合计",bg=DARK)
    ws2.cell(rI,2).fill=PatternFill("solid",fgColor=DARK); ws2.cell(rI,2).border=sd()
    c=ws2.cell(rI,3); c.value=s["income"]
    c.font=Font(name="Arial",bold=True,size=11,color=WHITE)
    c.fill=PatternFill("solid",fgColor=DARK)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.number_format='$#,##0.00'; c.border=sd()
    hdr(ws2.cell(rI,4),"",bg=DARK)

    ws3=wb.create_sheet("流水明细")
    for col,w in zip("ABCDEF",[8,8,52,12,22,10]): ws3.column_dimensions[col].width=w
    ws3.row_dimensions[1].height=22
    for col,txt in enumerate(["日期","来源","描述","金额","分类","状态"],1): hdr(ws3.cell(1,col),txt)
    for i,(_,row) in enumerate(df.iterrows()):
        r=i+2; ws3.row_dimensions[r].height=18; bg=LITE if i%2==0 else WHITE
        dat(ws3.cell(r,1),row["date"],bg=bg,align="center")
        dat(ws3.cell(r,2),row["src"],bg=bg,align="center")
        dat(ws3.cell(r,3),row["desc"],bg=bg)
        c=ws3.cell(r,4); c.value=row["amount"]
        c.font=Font(name="Arial",size=10,color="1A7A4A" if row["amount"]>0 else "000000")
        c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="right",vertical="center")
        c.number_format='$#,##0.00'; c.border=sd()
        dat(ws3.cell(r,5),row["cat"],bg=bg)
        sc={"auto":"1A7A4A","manual":"B85C00","pending":"C0392B"}
        dat(ws3.cell(r,6),row["status"],bg=bg,color=sc.get(row["status"],"000000"),align="center")

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ══════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════
st.set_page_config(page_title="ColorFour 财务分类", page_icon="💅", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500&family=DM+Serif+Display:ital@0;1&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif}
h1{font-family:'DM Serif Display',serif;font-weight:400;letter-spacing:-0.02em}
.mbox{background:#f7f7f5;border-radius:12px;padding:16px 20px;border:1px solid rgba(0,0,0,.06)}
.mlbl{font-size:11px;font-weight:500;letter-spacing:.08em;text-transform:uppercase;color:#888;margin-bottom:6px}
.mval{font-family:'DM Serif Display',serif;font-size:26px;font-weight:400}
.green{color:#1a7a4a}.red{color:#c0392b}
</style>
""", unsafe_allow_html=True)

if "df" not in st.session_state: st.session_state.df = None
if "month" not in st.session_state: st.session_state.month = "2026-02"
if "shopify" not in st.session_state: st.session_state.shopify = 0.0
if "bank_summary" not in st.session_state: st.session_state.bank_summary = {}

st.markdown("### ColorFour LLC")
st.markdown("# 财务流水*智能分类*")
st.caption("上传银行 Statement (PDF) 与 Payoneer 流水 (CSV)，自动分类并生成月度 P&L")
st.divider()

c1, c2 = st.columns([1,2])
with c1:
    month = st.selectbox("账单月份", [f"2026-{i:02d}" for i in range(1,13)], index=1)
    st.session_state.month = month
with c2:
    shopify = st.number_input("Shopify 净销售额（手动填入）",
                               min_value=0.0, value=st.session_state.shopify,
                               step=0.01, format="%.2f",
                               help="Shopify 收款不在银行流水里，请手动输入")
    st.session_state.shopify = shopify

c3, c4 = st.columns(2)
with c3:
    bank_file = st.file_uploader("🏦 银行 Statement", type=["pdf","csv"],
                                  help="自动识别 BofA / Chase")
with c4:
    pay_file  = st.file_uploader("💳 Payoneer 流水 CSV", type=["csv"])

rb, db = st.columns([1,5])
with rb: run = st.button("🚀 开始分类", type="primary", use_container_width=True)
with db: demo = st.button("演示数据（BofA 1月）")

if run:
    if not bank_file and not pay_file:
        st.warning("请至少上传一份文件")
    else:
        frames = []
        with st.spinner("解析中..."):
            if bank_file:
                try:
                    df_b, bname, bsummary = parse_bank_pdf(bank_file.read())
                    if not df_b.empty:
                        frames.append(df_b)
                        st.session_state.bank_summary = bsummary
                        st.success(f"✅ {bname} 解析完成，共 {len(df_b)} 笔")
                    else:
                        st.warning("银行流水解析为空，请检查 PDF 格式")
                except Exception as e:
                    st.error(f"银行解析失败：{e}")
            if pay_file:
                try:
                    df_p = parse_payoneer(pay_file.read())
                    if not df_p.empty:
                        frames.append(df_p)
                        st.success(f"✅ Payoneer 解析完成，共 {len(df_p)} 笔")
                    else:
                        st.warning("Payoneer CSV 解析为空")
                except Exception as e:
                    st.error(f"Payoneer 解析失败：{e}")
        if frames:
            st.session_state.df = pd.concat(frames, ignore_index=True)

if demo:
    demo_rows = [
        {"date":"01/02","desc":"TikTok Inc PAYMENT TikTok Shop payout 346879","amount":5625.40,"src":"BofA"},
        {"date":"01/06","desc":"TikTok Inc PAYMENT TikTok Shop payout 346889","amount":8866.50,"src":"BofA"},
        {"date":"01/07","desc":"TikTok Inc PAYMENT TikTok Shop payout 346903","amount":9195.92,"src":"BofA"},
        {"date":"01/08","desc":"TikTok Inc PAYMENT TikTok Shop payout 346913","amount":11800.33,"src":"BofA"},
        {"date":"01/12","desc":"TikTok Inc PAYMENT TikTok Shop payout 346918","amount":10346.21,"src":"BofA"},
        {"date":"01/13","desc":"TikTok Inc PAYMENT TikTok Shop payout 346934","amount":14338.99,"src":"BofA"},
        {"date":"01/21","desc":"TikTok Inc PAYMENT TikTok Shop payout 346965","amount":13233.48,"src":"BofA"},
        {"date":"01/02","desc":"PURCHASE 0101 TIKTOK ADS ADS.TIKTOK.COCA","amount":-5980.92,"src":"BofA"},
        {"date":"01/20","desc":"PURCHASE 0120 TIKTOK ADS ADS.TIKTOK.COCA","amount":-9968.86,"src":"BofA"},
        {"date":"01/02","desc":"Zelle payment to EVA SOLIS for Week52-53 payroll","amount":-1175.00,"src":"BofA"},
        {"date":"01/05","desc":"Zelle payment to NUTVIPA BUTRASIRT for Week52-53 payroll","amount":-623.40,"src":"BofA"},
        {"date":"01/13","desc":"ADP WAGE PAY ID:7670999705224TG","amount":-3982.81,"src":"BofA"},
        {"date":"01/14","desc":"ADP Tax ID:LT4TG 011401A01","amount":-1440.64,"src":"BofA"},
        {"date":"01/14","desc":"ADP PAY-BY-PAY ID:7670999705234TG","amount":-40.03,"src":"BofA"},
        {"date":"01/20","desc":"CHECKCARD Regus Management Group iwgplc.com TX","amount":-4035.38,"src":"BofA"},
        {"date":"01/05","desc":"TRANSFER COLORFOUR LLC:Mars Shipping Servic Confirmation# 4061732035","amount":-1470.15,"src":"BofA"},
        {"date":"01/07","desc":"TRANSFER COLORFOUR LLC:ColorFour LLC Confirmation# 1881149832","amount":-15000.00,"src":"BofA"},
        {"date":"01/09","desc":"Zelle payment to ZHENLI WANG","amount":-2781.00,"src":"BofA"},
        {"date":"01/26","desc":"PURCHASE 0124 KLAVIYO INC. SOFTWARE","amount":-100.00,"src":"BofA"},
        {"date":"01/02","desc":"CHECKCARD GOOGLE *Worksp Mountain View CA","amount":-39.51,"src":"BofA"},
        {"date":"01/09","desc":"PURCHASE 0109 FANTUAN DELIVE Fremont CA","amount":-41.95,"src":"BofA"},
        {"date":"01/06","desc":"External transfer fee - 3 Day","amount":-1.00,"src":"BofA"},
        {"date":"01/07","desc":"Check #135","amount":-960.00,"src":"BofA"},
        {"date":"02/10","desc":"Payment from COLORFOUR LLC - XX-445152","amount":20000.00,"src":"Payoneer"},
        {"date":"02/10","desc":"Payment to 招商银行 (9005)","amount":-7986.53,"src":"Payoneer"},
        {"date":"02/03","desc":"Payment to 中国银行 (9460)","amount":-2557.98,"src":"Payoneer"},
        {"date":"02/02","desc":"Payment to 中国银行 (1510)","amount":-1089.05,"src":"Payoneer"},
    ]
    df_demo = pd.DataFrame(demo_rows)
    df_demo["cat"]    = df_demo.apply(lambda r: cls_bank(r["desc"]) if r["src"]!="Payoneer" else cls_pay(r["desc"]), axis=1)
    df_demo["status"] = df_demo["cat"].apply(lambda c: "auto" if c!="待确认" else "pending")
    st.session_state.df = df_demo
    st.session_state.month = "2026-01"
    st.success(f"✅ 演示数据已加载，共 {len(df_demo)} 笔")

# ── Results ────────────────────────────────────────────────────
if st.session_state.df is not None:
    df = st.session_state.df.copy()
    s  = build_summary(df, st.session_state.shopify)
    pend = (df["status"]=="pending").sum()
    rate = round((df["status"]=="auto").sum()/len(df)*100)

    st.divider()

    # ── Reconciliation check against cover page ──────────────
    bsummary = st.session_state.get("bank_summary", {})
    if bsummary:
        bofa_df = df[df["src"]=="BofA"]
        parsed_dep = bofa_df[bofa_df["amount"]>0]["amount"].sum()
        parsed_wdl = abs(bofa_df[bofa_df["amount"]<0]["amount"].sum())
        cover_dep  = bsummary.get("deposits", None)
        cover_wdl  = bsummary.get("withdrawals", None)
        cover_chk  = bsummary.get("checks", 0)
        cover_fee  = bsummary.get("fees", 0)

        ok_dep = cover_dep is None or abs(parsed_dep - cover_dep) < 1.0
        total_wdl_cover = (cover_wdl or 0) + cover_chk + cover_fee
        ok_wdl = cover_wdl is None or abs(parsed_wdl - total_wdl_cover) < 1.0

        if ok_dep and ok_wdl:
            st.success(f"✅ 对账通过 — 存款 ${parsed_dep:,.2f} / 支出 ${parsed_wdl:,.2f} 与账单封面一致")
        else:
            msgs = []
            if not ok_dep and cover_dep:
                msgs.append(f"存款差异 ${abs(parsed_dep-cover_dep):,.2f}（解析 ${parsed_dep:,.2f} vs 封面 ${cover_dep:,.2f}）")
            if not ok_wdl and cover_wdl:
                msgs.append(f"支出差异 ${abs(parsed_wdl-total_wdl_cover):,.2f}（解析 ${parsed_wdl:,.2f} vs 封面 ${total_wdl_cover:,.2f}）")
            st.error("⚠️ 对账不符，可能有漏扫交易！\n" + " | ".join(msgs))

    m1,m2,m3,m4 = st.columns(4)
    for col, label, val, sub, clr in [
        (m1,"总收入",    f"${s['income']:,.2f}",  "TikTok + Shopify", "green"),
        (m2,"总支出",    f"${abs(s['expense']):,.2f}", "不含内部转账", ""),
        (m3,"净利润",    f"${s['net']:,.2f}",     f"利润率 {s['margin']}%", "green" if s['net']>0 else "red"),
        (m4,"自动分类率",f"{rate}%",              f"{pend} 笔待确认", ""),
    ]:
        with col:
            st.markdown(f'<div class="mbox"><div class="mlbl">{label}</div>'
                        f'<div class="mval {clr}">{val}</div>'
                        f'<div style="font-size:12px;color:#888;margin-top:4px">{sub}</div></div>',
                        unsafe_allow_html=True)

    st.markdown("")
    t1, t2, t3 = st.tabs(["📋 流水明细", "📊 P&L 汇总", "📈 图表"])

    with t1:
        if pend > 0:
            st.warning(f"⚠️ {pend} 笔交易待手动确认 — 已在下表中用🔴红色高亮")
        else:
            st.success("✅ 所有交易已分类完成")

        fa,fb,fc = st.columns(3)
        with fa: cf = st.selectbox("分类", ["全部"]+ALL_CATS)
        with fb: sf = st.selectbox("来源", ["全部","BofA","Chase","Payoneer"])
        with fc: tf = st.selectbox("状态", ["全部","auto","pending","manual"])

        filt = df.copy()
        if cf!="全部": filt=filt[filt["cat"]==cf]
        if sf!="全部": filt=filt[filt["src"]==sf]
        if tf!="全部": filt=filt[filt["status"]==tf]

        # ── Highlighted view ─────────────────────────────────
        display_df = filt[["date","src","desc","amount","cat","status"]].rename(columns={
            "date":"日期","src":"来源","desc":"描述","amount":"金额","cat":"分类","status":"状态"})

        def highlight_pending(row):
            if row["状态"] == "pending":
                return ["background-color:#fde8e8; color:#c0392b; font-weight:500"] * len(row)
            return [""] * len(row)

        styled = display_df.style.apply(highlight_pending, axis=1).format({"金额": "${:,.2f}"})
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # ── Editable correction (FIX: use key + save button) ─
        st.markdown("**✏️ 修改分类（改完后点击「保存修改」）**")

        edit_df = filt[["date","src","desc","amount","cat","status"]].copy()
        edit_df = edit_df.rename(columns={
            "date":"日期","src":"来源","desc":"描述","amount":"金额","cat":"分类","status":"状态"})
        # Store the original index mapping so we can write back
        edit_df["_orig_idx"] = filt.index.tolist()

        edited = st.data_editor(
            edit_df.drop(columns=["_orig_idx"]),
            column_config={
                "分类": st.column_config.SelectboxColumn("分类", options=ALL_CATS, width="medium"),
                "金额": st.column_config.NumberColumn("金额", format="$%.2f"),
                "日期": st.column_config.TextColumn("日期", disabled=True),
                "来源": st.column_config.TextColumn("来源", disabled=True),
                "描述": st.column_config.TextColumn("描述", disabled=True),
                "状态": st.column_config.TextColumn("状态", disabled=True),
            },
            use_container_width=True, hide_index=True, num_rows="fixed",
            key="cat_editor",
        )

        if st.button("💾 保存修改", type="primary"):
            changed = 0
            for i, orig_idx in enumerate(edit_df["_orig_idx"]):
                nc = edited.iloc[i]["分类"]
                if nc != st.session_state.df.at[orig_idx, "cat"]:
                    st.session_state.df.at[orig_idx, "cat"]    = nc
                    st.session_state.df.at[orig_idx, "status"] = "manual"
                    changed += 1
            if changed:
                st.success(f"✅ 已保存 {changed} 条分类修改")
                st.rerun()
            else:
                st.info("没有检测到修改")

    with t2:
        pl1, pl2 = st.columns(2)
        with pl1:
            for label, rows_f, total, clr in [
                ("收入", [(k,v) for k,v in s["totals"].items() if v>0 and k not in EXCLUDE_PL], s["income"], "#1a7a4a"),
                ("支出", [(k,v) for k,v in s["totals"].items() if v<0 and k not in EXCLUDE_PL], s["expense"], None),
            ]:
                st.markdown(f"**{label}**")
                for cat, amt in sorted(rows_f, key=lambda x:-abs(x[1])):
                    color = "#1a7a4a" if amt>0 else "inherit"
                    st.markdown(f"<div style='display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid #f0f0ed;font-size:13px'>"
                                f"<span>{cat}</span><span style='color:{color}'>${abs(amt):,.2f}</span></div>",
                                unsafe_allow_html=True)
                col_color = clr if clr else "#1a1a1a"
                st.markdown(f"<div style='display:flex;justify-content:space-between;padding:8px 0;font-weight:500'>"
                            f"<span>{label}合计</span><span style='color:{col_color}'>${abs(total):,.2f}</span></div>",
                            unsafe_allow_html=True)
                st.markdown("")
            nc = "#1a7a4a" if s["net"]>0 else "#c0392b"
            st.markdown(f"<div style='display:flex;justify-content:space-between;padding:14px 0;"
                        f"font-family:\"DM Serif Display\",serif;font-size:22px;border-top:2px solid #1a1a1a'>"
                        f"<span>净利润</span><span style='color:{nc}'>${s['net']:,.2f}</span></div>",
                        unsafe_allow_html=True)
        with pl2:
            exp_d = [(k,abs(v)) for k,v in s["totals"].items() if v<0 and k not in EXCLUDE_PL]
            if exp_d:
                exp_d.sort(key=lambda x:-x[1])
                top8=exp_d[:8]; others=sum(x[1] for x in exp_d[8:])
                if others>0: top8.append(("其他",others))
                lbls,vals=zip(*top8)
                fig=go.Figure(go.Pie(labels=lbls,values=vals,hole=0.45,textinfo="percent"))
                fig.update_layout(margin=dict(l=0,r=0,t=0,b=0),height=320,
                                  paper_bgcolor="rgba(0,0,0,0)",font_family="DM Sans")
                st.plotly_chart(fig, use_container_width=True)

    with t3:
        daily = df.groupby("date")["amount"].sum().reset_index()
        fig2 = px.bar(daily, x="date", y="amount",
                      color=daily["amount"].apply(lambda x:"收入" if x>0 else "支出"),
                      color_discrete_map={"收入":"#1a7a4a","支出":"#e74c3c"},
                      title="每日收支")
        fig2.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
                           height=280,margin=dict(l=0,r=0,t=40,b=0))
        st.plotly_chart(fig2, use_container_width=True)

        cat_t = df.groupby("cat")["amount"].sum().reset_index()
        cat_t = cat_t[~cat_t["cat"].isin(EXCLUDE_PL)].sort_values("amount")
        fig3 = px.bar(cat_t, x="amount", y="cat", orientation="h",
                      color=cat_t["amount"].apply(lambda x:"收入" if x>0 else "支出"),
                      color_discrete_map={"收入":"#1a7a4a","支出":"#2e75b6"},
                      title="各类别汇总")
        fig3.update_layout(paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
                           showlegend=False,height=max(300,len(cat_t)*22),
                           margin=dict(l=0,r=0,t=40,b=0))
        st.plotly_chart(fig3, use_container_width=True)

    st.divider()
    st.markdown("#### 导出")
    e1,e2,e3 = st.columns(3)
    with e1:
        st.download_button("📄 财报 CSV", data=export_csv(df,month,shopify),
                           file_name=f"ColorFour_{month}_财务报表.csv", mime="text/csv",
                           use_container_width=True)
    with e2:
        st.download_button("📊 财报 Excel", data=export_excel(df,month,shopify),
                           file_name=f"ColorFour_{month}_财务报表.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
    with e3:
        st.download_button("📋 原始流水", data=df.to_csv(index=False).encode("utf-8-sig"),
                           file_name=f"ColorFour_{month}_原始流水.csv", mime="text/csv",
                           use_container_width=True)
